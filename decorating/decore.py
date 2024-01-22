import os
import time

import openpyxl
import win32com.client
from openpyxl.chart import BarChart, Reference, DoughnutChart, LineChart
from openpyxl.chart.label import DataLabelList, DataLabel
from openpyxl.chart.marker import DataPoint
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.text import RichText
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties, Font as txFont
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.table import Table, TableStyleInfo
import pandas as pd
from openpyxl.xml.functions import fromstring

from logs import log

class OpenWorkbookException(Exception):
    def __str__(self):
        return 'Ошибка при открытии документа'

class ChartParams:
    def __init__(self, width, height, left, top, num_chart, num_slide):
        self.top = top
        self.left = left
        self.height = height
        self.width = width
        self.num_chart = num_chart
        self.num_slide = num_slide

CHART_PARAMS = {
    'Количество документов от ИФНС в разрезе периодов': ChartParams(900, 450, 30, 60, 1, 3),
    'Анализ по видам документов от ИФНС': ChartParams(440, 450, 30, 60, 1, 4),
    'Структурный анализ количества документов в разрезе организаций': ChartParams(440, 450, 500, 50, 2, 4),
    'Структурный анализ проверяемых деклараций в рамках камеральных проверок': ChartParams(440, 450, 30, 60, 1, 5),
    'Анализ в разрезе налоговых проверок': ChartParams(440, 450, 500, 50, 2, 5),
    'Анализ в разрезе подрядчиков': ChartParams(440, 450, 30, 60, 2, 6),
    'Динамика количества проверяемых подрядчиков на фоне темпов роста количества требований': ChartParams(440, 450, 500, 50, 3, 6)
}

class TextBoxParams:
    def __init__(self, left, top, width, height, dim):
        self.top = top
        self.left = left
        self.height = height
        self.width = width
        self.dim = dim

START_COLUMN, INDENT, START_ROW = 2, 4, 2
DATE_COLUMN = ['Дата', 'Дата документа', 'Получен']
HEADERS_FONT = Font(
        name='Calibri',
        size=16,
        bold=True,
        italic=False,
        vertAlign=None,
        underline='none',
        strike=False,
        color='FF000000'
            )

TIME_WAITING = 2
SHEETNAME = 'Сверка'
xml = f"""
<txPr>
  <a:p xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
    <a:r>
       <a:rPr b="1" i="0" sz="1000" spc="-1" strike="noStrike">
          <a:solidFill>
             <a:srgbClr val="499EFA" />
          </a:solidFill>
          <a:latin typeface="Calibri" />
       </a:rPr>
       <a:t>Наименование</a:t>
    </a:r>
  </a:p>
</txPr>
"""

TEMPLATE_FILENAME = r'templates\Шаблон.pptx'
TEMPLATE_FILENAME1 = r'templates\Шаблон1.pptx'
PRES_FILENAME = 'Отчет по незакрытым авансам.pptx'
TAX_PRES_FILENAME = 'Отчет по налогам.pptx'
TXT_BOX_PARAMS = {
    'Сумма незакрытого аванса': TextBoxParams(50, -100, 400, 450, 'млрд руб.'),
    'Сумма НДС с незакрытого аванса': TextBoxParams(500, -100, 400, 450, 'млн руб.'),
    'Количество непредоставленных документов': TextBoxParams(50, 90, 400, 450, 'шт'),
    'Сумма авансирования': TextBoxParams(500, 90, 400, 450, 'млрд руб.')
}

TXT_BOX_PARAMS_FOR_TAXES = {
    'Доля квитанций отправленных вовремя': TextBoxParams(50, -100, 400, 450, '%'),
    'Доля ответов на Требования, отправленных вовремя': TextBoxParams(500, -100, 400, 450, '%'),
    'Количество требований, отправленных позже срока': TextBoxParams(50, 90, 400, 450, 'Требований'),
    'Прогноз начисления штрафа': TextBoxParams(500, 90, 400, 450, 'руб.')
}

def split_dict_report(report_dict: dict) -> tuple:
    adv_dict = {k:v for k, v in report_dict.items() if k.startswith('REPORT')}
    err_dict = {k:v for k, v in report_dict.items() if not k.startswith('REPORT')}
    return err_dict, adv_dict
def create_result_files(report_dict: dict, err_name: str=None, adv_name: str=None, type_report: str=None) -> None:
    log.info('Оформление результирующих файлов')
    if type_report == 'all':
        err_dict, adv_dict = split_dict_report(report_dict)
        decorating_excel_list(err_dict, err_name)
        decorating_excel_list(adv_dict, adv_name, type_report)
    elif type_report == 'only_errors':
        decorating_excel_list(report_dict, err_name)
    else:
        decorating_excel_list(report_dict, adv_name, type_report)


def decorating_excel_list(report_dict: dict, savename: str, type_report: str=None) -> None:
    log.info(f'Оформление документа следующего типа {type_report}')
    wb = openpyxl.Workbook()
    sheet = wb.get_sheet_by_name(wb.sheetnames[0])
    sheet.title = SHEETNAME
    row = 3
    col = 7
    for title, res in report_dict.items():
        add_smart_table(wb, sheet, res, title)
        if type_report == None:
            add_charts_for_tax_report(wb, sheet, title, row, col, len(res))
            row+=len(res)+3
    if type_report != None:
        set_max_width(report_dict, sheet)
    else:
        pass
    if type_report in ('all', 'only_advances'):
        add_bar_chart(wb, sheet,'REPORT_ТОП10'.replace(' ', '_'), 'G11', 'ТОП-10 подрядчиков по общей сумме незакрытого аванса', 'Наименование подрядчиков', 'Сумма (млн руб.)')
        add_dognut_chart(wb, sheet,'REPORT_Анализ'.replace(' ', '_'), 'G25', 'Структурный анализ НДС по документам')
        add_hist(wb, sheet, 'REPORT_Средняя разница'.replace(' ', '_'), 'G40', len(report_dict['REPORT_Средняя разница']), 'Среднее количество дней непредоставления документов')
    wb.save(savename)
    wb.close()
    if type_report in ('all', 'only_advances'):
        create_presentation(savename, report_dict['REPORT_Сумма НДС'])
    elif type_report == None:
        best_create_pres(savename, TEMPLATE_FILENAME1, report_dict['Информативный блок'])

def add_charts_for_tax_report(wb, sheet, title, row, col, length):
    AXIS_DICT = {
        'Поквартальный анализ полученных документов': ('Период', 'Количество документов, шт.', 'Количество документов от ИФНС в разрезе периодов'),
        'Анализ в разрезе видов документов': ('Количество шт.', 'Вид документа', "Анализ по видам документов от ИФНС"),
        'Анализ по виду организаций': (None, None, "Структурный анализ количества документов в разрезе организаций",),
        'Анализ по виду налога': (None, None, "Структурный анализ проверяемых деклараций в рамках камеральных проверок"),
        'Анализ по виду проверки': ('Количество шт.', 'Вид налоговой проверки', "Анализ в разрезе налоговых проверок"),
        'Основные подрядчики': ('Количество шт.', 'Наименование подрядчика', "Анализ в разрезе подрядчиков"),
        'Анализ требований': ('Год', 'Количество требований', 'Количество подрядчиков', "Динамика количества проверяемых подрядчиков на фоне темпов роста количества требований"),

    }
    rng = f'{get_column_letter(col)}{row}'
    if title in ('Анализ в разрезе видов документов', 'Анализ по виду проверки', 'Основные подрядчики'):
        y_title, x_title, title_for_excel = AXIS_DICT[title]
        add_bar_chart(wb, sheet,title.replace(' ', '_'), rng, title_for_excel, x_title, y_title, lbl_pos='inBase')
    if title in ('Поквартальный анализ полученных документов'):
        x_title, y_title, title_for_excel = AXIS_DICT[title]
        add_bar_chart(wb, sheet, title.replace(' ', '_'), rng, title_for_excel, x_title, y_title,type_chart='col', lbl_pos='outEnd')
    if title in ('Анализ по виду налога', 'Анализ по виду организаций'):
        _, _, title_for_excel = AXIS_DICT[title]
        clr_list = ("DFE5EF","499EFA")
        if title == 'Анализ по виду организаций':
            clr_list = create_color_list(length)
        add_dognut_chart(wb, sheet, title.replace(' ', '_'), rng, title_for_excel,dp=length, clr_list=clr_list)
    if title == 'Анализ требований':
        x_title, y1_title, y2_title, title_for_excel = AXIS_DICT[title]
        add_hist(wb, sheet, title.replace(' ', '_'), rng, length, y1_title, y2_title, title_for_excel, x_title)


def create_color_list(length: int) -> tuple:
    clr_list = ["DFE5EF","499EFA"]
    green = 123
    blue = 252
    red = 0
    for _ in range(length):
        if green >= 255:
            green = 0
        elif blue<=0:
            blue = 252
        elif red >= 252:
            red = 0
        green+=10
        blue-=10
        red+=10
        clr = rgb2hex(red, green, blue)
        clr_list.append(clr)
    return tuple(clr_list)

def rgb2hex(r,g,b):
    return "{:02x}{:02x}{:02x}".format(r,g,b)
def add_smart_table(wb, ws,  result: pd.DataFrame, title: str) -> None:
    log.info(f'Оформление таблицы {title}')
    global START_ROW
    if title in ('REPORT_Перегрупировка', 'REPORT_Все контрагенты'):
        START_ROW = 2
        ws = wb.create_sheet(title.replace('REPORT_', '').replace(' ', '_'))
    result = edit_date_columns(result)
    title = title.replace('REPORT_', '')
    ws.cell(row=START_ROW, column=START_COLUMN).value = title
    ws.cell(row=START_ROW, column=START_COLUMN).font = HEADERS_FONT
    length = len(result)
    if not result.empty:
        rows = dataframe_to_rows(result, index=False)
        for r_idx, row in enumerate(rows, START_ROW + 1):
            for c_idx, value in enumerate(row, START_COLUMN):
                if title in ('Анализ по виду организаций', 'Анализ по виду налога', 'Анализ') and c_idx==3:
                    ws.cell(row=r_idx, column=c_idx, value=value).number_format = '0%'
                else:
                    ws.cell(row=r_idx, column=c_idx, value=value)

        length = len(result)
        col_length = len(result.columns) - 1
        table_range = f'{get_column_letter(START_COLUMN)}{START_ROW + 1}:{get_column_letter(START_COLUMN + col_length)}{START_ROW + length + 1}'
        title_wout_space = title.replace(' ', '_')
        table = Table(displayName=title_wout_space, ref=table_range)
        style = TableStyleInfo(name="TableStyleMedium2", showFirstColumn=False,
                               showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        table.tableStyleInfo = style
        ws.add_table(table)
        if title in ('Средняя разница', 'Анализ требований'):
            data_range_avg = f'{get_column_letter(START_COLUMN + 1)}{START_ROW + 1}:{get_column_letter(START_COLUMN + col_length-1)}{START_ROW + length + 1}'
            data_range_sum = f'{get_column_letter(START_COLUMN + 2)}{START_ROW + 1}:{get_column_letter(START_COLUMN + col_length)}{START_ROW + length + 1}'
            wb.defined_names.add(create_defined_range(data_range_avg, title_wout_space, 'data_avg'))
            wb.defined_names.add(create_defined_range(data_range_sum, title_wout_space, 'data_sum'))
        else:
            data_range = f'{get_column_letter(START_COLUMN+1)}{START_ROW + 1}:{get_column_letter(START_COLUMN + col_length)}{START_ROW + length + 1}'
            wb.defined_names.add(create_defined_range(data_range, title_wout_space, 'data'))
        category_range = f'{get_column_letter(START_COLUMN)}{START_ROW + 2}:{get_column_letter(START_COLUMN)}{START_ROW + length + 1}'
        wb.defined_names.add(create_defined_range(category_range, title_wout_space, 'cat'))
        if title in ('Перегрупировка', 'Все контрагенты'):
            width_dict = full_width_dict(dict(), result)
            for column, width in width_dict.items():
                ws.column_dimensions[get_column_letter(column)].width = width
    START_ROW += length + INDENT

def create_defined_range(rng: str,  title: str, type_rng: str):
    def_rng = DefinedName(f'{type_rng}_{title}', attr_text=f'{SHEETNAME}!{rng}')
    return def_rng



def edit_date_columns(table: pd.DataFrame) -> pd.DataFrame:
    for col in DATE_COLUMN:
        if col in table.columns:
            table[col] = table[col].apply(str)
    return table

def full_width_dict(width_dict: dict, report: pd.DataFrame) -> dict:
    for i, column in enumerate(report.columns, START_COLUMN):
        data_list = [len(str(value)) for value in report[column]]
        data_list.append(len(column))
        max_width = max(data_list) + INDENT
        current_dict = width_dict.get(i, 1)
        if max_width > current_dict:
            width_dict[i] = max_width
    return width_dict
def set_max_width(report_dict: dict, ws, opt:bool=True) -> None:
    width_dict = dict()
    for title, report in report_dict.items():
        if title not in ('REPORT_Перегрупировка', 'REPORT_Все контрагенты'):
            width_dict = full_width_dict(width_dict, report)
    for column, width in width_dict.items():
        ws.column_dimensions[get_column_letter(column)].width = width

def add_dognut_chart(wb, ws, table_range: str, past_range: str, title: str, dp: int=2, clr_list: tuple=("DFE5EF","499EFA")) -> None:
    log.info(f'Создание круговой диаграммы')
    table_range = table_range.replace('REPORT_', '')
    data_range = '!'.join(list(wb.defined_names[f'data_{table_range}'].destinations)[0])
    cat_range = '!'.join(list(wb.defined_names[f'cat_{table_range}'].destinations)[0])
    data = Reference(ws, range_string=data_range)
    cats = Reference(ws, range_string=cat_range)
    chart = DoughnutChart(holeSize=50)
    chart.style = 1
    chart.legend.position = 'b'
    chart.add_data(data, titles_from_data=True)
    slices = [DataPoint(idx=i) for i in range(dp)]
    color_list = clr_list
    for idx, point in enumerate(slices):
        col_idx = idx % len(color_list)
        point.graphicalProperties.solidFill = color_list[col_idx]
        point.graphicalProperties.ln.solidFill = color_list[col_idx]
    chart.series[0].data_points = slices
    chart.set_categories(cats)
    chart.series[0].dLbls = DataLabelList()
    chart.series[0].dLbls.showVal = 1
    for i in range(dp):
        if i == 0:
            set_data_labels_setting(chart, '499EFA', i)
        else:
            set_data_labels_setting(chart, 'DFE5EF', i)
    set_axis_setting(chart, title, '000000', 'title', 0, 1)
    ws.add_chart(chart, past_range)
def add_hist(wb, ws, table_range: str, past_range: str, length: int, y1_title:str, y2_title:str=None, title:str=None,x_title:str=None) -> None:
    log.info(f'Создание вертикальной гистограммы')
    table_range = table_range.replace('REPORT_', '')
    data_range = '!'.join(list(wb.defined_names[f'data_avg_{table_range}'].destinations)[0])
    cat_range = '!'.join(list(wb.defined_names[f'cat_{table_range}'].destinations)[0])
    bar_chart = BarChart()
    data = Reference(ws, range_string=data_range)
    cats = Reference(ws, range_string=cat_range)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    set_color(bar_chart, 0, '499EFA')
    bar_chart.dLbls = DataLabelList()
    bar_chart.dLbls.showVal = 1
    bar_chart.dLbls.dLblPos = 'inBase'
    bar_chart.x_axis.majorGridlines = None
    bar_chart.y_axis.majorGridlines = None
    bar_chart.y_axis.title = y1_title
    set_axis_setting(bar_chart, y1_title, '000000', 'y', 1, 1)
    # if y2_title != None:
    #     set_axis_setting(bar_chart, title, '000000', 'title', 0, 1)
    for i in range(length):
        set_data_labels_setting(bar_chart, 'FFFFFF', i, False)
    line_chart = LineChart()
    data_range = '!'.join(list(wb.defined_names[f'data_sum_{table_range}'].destinations)[0])
    line_chart.add_data(data_range, titles_from_data=True)
    line_chart.x_axis.majorGridlines = None
    line_chart.y_axis.majorGridlines = None
    set_color(line_chart, 0, '000000')
    line_chart.series[0].graphicalProperties.line.width = 25000
    line_chart.dLbls = DataLabelList()
    line_chart.dLbls.showVal = 1
    line_chart.y_axis.axId = 200
    line_chart.y_axis.crosses = "max"
    if y2_title != None:
        line_chart.y_axis.title = y2_title
        line_chart.x_axis.title = x_title
        set_axis_setting(line_chart, x_title, '000000', 'x', 1, 1)
        set_axis_setting(line_chart, title, '000000', 'title', 0, 1)
    line_chart+=bar_chart
    line_chart.legend.position = 'b'
    ws.add_chart(line_chart, past_range)
def add_bar_chart(wb, ws, table_range: str, past_range: str, title: str, x_title: str, y_title: str, type_chart: str = 'bar', lbl_pos: str='inEnd') -> None:
    log.info(f'Создание горизонтальной гистограммы')
    # x_title, y_title = 'Наименование подрядчиков', 'Сумма (млн руб.)'
    table_range = table_range.replace('REPORT_', '')
    data_range = '!'.join(list(wb.defined_names[f'data_{table_range}'].destinations)[0])
    cat_range = '!'.join(list(wb.defined_names[f'cat_{table_range}'].destinations)[0])
    bar_chart = BarChart()
    bar_chart.type = type_chart
    set_axis_setting(bar_chart,title, '000000', 'title', 0, 1)
    data = Reference(ws, range_string=data_range)
    cats = Reference(ws, range_string=cat_range)
    bar_chart.add_data(data, titles_from_data=True)
    bar_chart.set_categories(cats)
    bar_chart.style = 1
    if type_chart != 'col':
        bar_chart.grouping = "stacked"
    bar_chart.overlap = 100
    bar_chart.dLbls = DataLabelList()
    bar_chart.dLbls.showVal = 1
    bar_chart.dLbls.dLblPos = lbl_pos

    bar_chart.x_axis.title = x_title
    bar_chart.y_axis.title = y_title
    set_axis_setting(bar_chart, x_title, '499EFA', 'x', 1, 1)
    set_axis_setting(bar_chart, y_title, '499EFA', 'y', 1, 1)
    set_color(bar_chart, 0, '499EFA')
    if title == 'ТОП-10 подрядчиков по общей сумме незакрытого аванса':
        set_color(bar_chart, 1, 'DFE5EF')

    bar_chart.legend.position = 'b'
    bar_chart.x_axis.majorGridlines = None
    bar_chart.y_axis.majorGridlines = None

    ws.add_chart(bar_chart, past_range)


def set_color(chart, number: int,  color: str):
    chart.series[number].graphicalProperties.line.solidFill = color
    chart.series[number].graphicalProperties.solidFill = color

def set_axis_setting(chart, title, color, axis: str, bold:int=1, italic:int=0) -> None:
    xml = f"""
    <txPr>
      <a:p xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
        <a:r>
           <a:rPr b="{bold}" i="{italic}" sz="1000" spc="-1" strike="noStrike">
              <a:solidFill>
                 <a:srgbClr val="{color}" />
              </a:solidFill>
              <a:latin typeface="Calibri" />
           </a:rPr>
           <a:t>{title}</a:t>
        </a:r>
      </a:p>
    </txPr>
    """
    if axis == 'y':
        chart.y_axis.title.tx.rich = RichText.from_tree(fromstring(xml))
    elif axis == 'x':
        chart.x_axis.title.tx.rich = RichText.from_tree(fromstring(xml))
    elif axis == 'title':
        chart.title = title
        chart.title.tx.rich = RichText.from_tree(fromstring(xml))



def set_data_labels_setting(chart, color:str, idx: int, opt:bool=True) -> None:
    xml = f"""
    <txPr>
      <a:bodyPr wrap="square" lIns="38100" tIns="19050" rIns="38100" bIns="19050" anchor="ctr" xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
        <a:spAutoFit />
      </a:bodyPr>
      <a:lstStyle xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main" />
      <a:p xmlns:a="http://schemas.openxmlformats.org/drawingml/2006/main">
        <a:pPr >
          <a:defRPr b="1" i="1" sz="1500" spc="-1" strike="noStrike">
            <a:solidFill>
              <a:srgbClr val="{color}" />
            </a:solidFill>
          </a:defRPr>
        </a:pPr>
        <a:endParaRPr lang="de-DE" />
      </a:p>
    </txPr>
        """
    text_props = RichText.from_tree(fromstring(xml))
    label = DataLabel(idx=idx, txPr=text_props)
    if opt:
        chart.series[0].dLbls.dLbl.append(label)
    else:
        chart.dLbls.dLbl.append(label)

def create_pres_path(filename: str, name:str=PRES_FILENAME) -> str:
    path = filename[:filename.rfind('\\')]
    path = fr'"{path}/{name}"'
    return path
def create_presentation(filename, describe_data: pd.DataFrame) -> None:
    log.info(f'Создание презентации')
    savepath = create_pres_path(filename)
    PowerPoint = win32com.client.Dispatch("PowerPoint.Application")
    Excel = win32com.client.Dispatch("Excel.Application")
    try:
        presentation = PowerPoint.Presentations.Open(os.path.abspath(TEMPLATE_FILENAME), WithWindow=1)
        workbook = Excel.Workbooks.Open(Filename=filename, ReadOnly=1, UpdateLinks=False)
        time.sleep(5)
    except Exception as exp:
        log.info('Ошибка при открытии документа')
        log.exception(exp)
        Excel.Quit()
        PowerPoint.Quit()
        raise OpenWorkbookException
    try:
        for ws in workbook.Worksheets:
            if list(ws.ChartObjects()) != []:
                bar_chart, dognut_chart, hist_chart = ws.ChartObjects()[0], ws.ChartObjects()[1], ws.ChartObjects()[2]
                bar_chart.Activate()
                bar_chart.Copy()
                presentation.Slides.Item(3).Select()
                slide = presentation.Slides.Item(3)
                PowerPoint.CommandBars.ExecuteMso("PasteSourceFormatting")
                time.sleep(TIME_WAITING)
                set_chart_parametres(slide, 1, 440, 450, 30, 60)
                dognut_chart.copy()
                PowerPoint.CommandBars.ExecuteMso("PasteSourceFormatting")
                time.sleep(TIME_WAITING)
                set_chart_parametres(slide, 2, 440, 450, 500, 50)
                presentation.Slides.Item(2).Select()
                slide = presentation.Slides.Item(2)
                for k, v in TXT_BOX_PARAMS.items():
                    text = describe_data[describe_data['Тип']==k]['Сумма'].iloc[0]
                    if k == 'Количество непредоставленных документов':
                        text = int(text)
                    text = f'{text} {v.dim}'
                    add_text_to_slide(slide, text, 1, v.left, v.top, v.width, v.height)
                time.sleep(TIME_WAITING)
                presentation.Slides.Item(4).Select()
                slide = presentation.Slides.Item(4)
                hist_chart.copy()
                PowerPoint.CommandBars.ExecuteMso("PasteSourceFormatting")
                time.sleep(TIME_WAITING)
                set_chart_parametres(slide, 2, 900, 450, 30, 60)
                time.sleep(TIME_WAITING)
    except Exception as exp:
        log.info('Ошибка при оформлении презентации')
        log.exception(exp)
    finally:
        presentation.SaveAs(savepath)
        presentation.Close()
        workbook.Close()
        Excel.Quit()
        PowerPoint.Quit()



def set_chart_parametres(slide, num_shape:int, width:int, height:int, left:int, top:int) -> None:
    shape = slide.Shapes[num_shape]
    shape.Width = width
    shape.Height = height
    shape.Left = left
    shape.Top = top
def add_text_to_slide(slide, text, orient:int, left: int, top:int, width:int, height:int) -> None:
    textbox = slide.Shapes.AddTextbox(orient, left, top, width, height)
    textbox.TextFrame.TextRange.Text = text
    textbox.TextFrame.TextRange.Font.Bold = True
    textbox.TextFrame.TextRange.Font.Color = rgbToInt((0, 125, 255))


def rgbToInt(rgb: tuple) ->int:
    colorInt = rgb[0] + (rgb[1] * 256) + (rgb[2] * 256 * 256)
    return colorInt


def best_create_pres(filename: str, template_name: str, describe_data: pd.DataFrame) -> None:
    log.info(f'Создание презентации')
    savepath = create_pres_path(filename, TAX_PRES_FILENAME)
    PowerPoint = win32com.client.Dispatch("PowerPoint.Application")
    Excel = win32com.client.Dispatch("Excel.Application")
    try:
        presentation = PowerPoint.Presentations.Open(os.path.abspath(template_name), WithWindow=1)
        workbook = Excel.Workbooks.Open(Filename=filename, ReadOnly=1, UpdateLinks=False)
        time.sleep(5)
    except Exception as exp:
        log.info('Ошибка при открытии документа')
        log.exception(exp)
        Excel.Quit()
        PowerPoint.Quit()
        raise OpenWorkbookException
    try:
        for ws in workbook.Worksheets:
            if list(ws.ChartObjects()) != []:
                for chart in list(ws.ChartObjects()):
                    title = chart.Chart.ChartTitle.Text
                    params = CHART_PARAMS[title]
                    chart.Activate()
                    chart.Copy()
                    presentation.Slides.Item(params.num_slide).Select()
                    slide = presentation.Slides.Item(params.num_slide)
                    PowerPoint.CommandBars.ExecuteMso("PasteSourceFormatting")
                    time.sleep(TIME_WAITING)
                    set_chart_parametres(slide, params.num_chart,  params.width, params.height, params.left, params.top)
        presentation.Slides.Item(2).Select()
        slide = presentation.Slides.Item(2)
        for k, v in TXT_BOX_PARAMS_FOR_TAXES.items():
                text = describe_data[describe_data['Тип']==k]['Сумма'].iloc[0]
                text = f'{text} {v.dim}'
                add_text_to_slide(slide, text, 1, v.left, v.top, v.width, v.height)
    except Exception as exp:
        log.info('Ошибка при оформлении презентации')
        log.exception(exp)
    finally:
        presentation.SaveAs(savepath)
        time.sleep(TIME_WAITING)
        presentation.Close()
        time.sleep(TIME_WAITING)
        workbook.Close()
        time.sleep(TIME_WAITING)
        Excel.Quit()
        PowerPoint.Quit()